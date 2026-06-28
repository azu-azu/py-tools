from __future__ import annotations

from difflib import get_close_matches

from openpyxl.workbook import Workbook


class SheetNotFoundError(Exception):
    """期待するシートが workbook に存在しない場合のエラー。"""


def resolve_sheet(wb: Workbook, expected: str):
    """expected 名のシートを返す。見つからなければ fuzzy 候補を提示して raise。"""
    if expected in wb.sheetnames:
        return wb[expected]

    candidates = get_close_matches(expected, wb.sheetnames, n=3, cutoff=0.5)
    sheets_list = ", ".join(wb.sheetnames)

    if candidates:
        suggestion = ", ".join(candidates)
        raise SheetNotFoundError(
            f"シート '{expected}' が見つかりません。"
            f"\n  typo の可能性: {suggestion}"
            f"\n  → Excel でシート名を '{expected}' に修正してください。"
            f"\n  (全シート: {sheets_list})"
        )
    raise SheetNotFoundError(
        f"シート '{expected}' が見つかりません。"
        f"\n  全シート: {sheets_list}"
        f"\n  → Excel に '{expected}' シートを作成してください。"
    )
